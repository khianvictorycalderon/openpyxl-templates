import openpyxl

import openpyxl

def opxl_delete(file_path, sheet_name, condition=None, row=None):
    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in workbook.")
    
    ws = wb[sheet_name]

    # Validate inputs
    if row is not None and condition is not None:
        raise ValueError("Provide only one of 'row' or 'condition'.")

    if row is not None:
        # Since header is row 1, data starts at row 2 in Excel
        excel_row_to_delete = row + 1
        max_data_row = ws.max_row - 1  # number of data rows excluding header

        if row < 1 or row > max_data_row:
            raise ValueError(f"Data row {row} is out of range. Max data rows: {max_data_row}")

        ws.delete_rows(excel_row_to_delete)
        wb.save(file_path)
        print(f"Row {row} successfully deleted.")
        return

    if condition is not None:
        headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

        for key in condition.keys():
            if key not in headers:
                raise ValueError(f"Column '{key}' not found in sheet headers.")

        rows_to_delete = []
        for row_idx in range(2, ws.max_row + 1):
            match = True
            for col_name, cond_val in condition.items():
                cell_value = ws.cell(row=row_idx, column=headers[col_name]).value
                if callable(cond_val):
                    if not cond_val(cell_value):
                        match = False
                        break
                else:
                    if cell_value != cond_val:
                        match = False
                        break
            if match:
                rows_to_delete.append(row_idx)

        if not rows_to_delete:
            print("No rows matched the condition. Nothing deleted.")
        else:
            for r in reversed(rows_to_delete):
                ws.delete_rows(r)
            wb.save(file_path)
            print(f"Deleted {len(rows_to_delete)} row(s) matching the condition.")
        return

    # If neither row nor condition provided, delete all rows except header
    if row is None and condition is None:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            wb.save(file_path)
            print(f"Deleted all rows except the header.")
        else:
            print("No data rows to delete.")
            

# Sample Usage

db = "sample.xlsx"
sheet = "Sample_Sheet"

# 3 ways to delete a row:

# Delete row 1
opxl_delete(db, sheet, row=1)

# Delete all rows where First Name == "Karen"
opxl_delete(db, sheet, condition={"First Name": "Karen"})

# Delete all rows where Age >= 35
opxl_delete(db, sheet, condition={"Age": lambda x: x is not None and x >= 35})

# Simply delete all rows (Except the headers ofcourse)
opxl_delete(db, sheet)