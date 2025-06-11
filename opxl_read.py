import openpyxl

def opxl_read(file_path, sheet_name, columns=None, condition=None):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    except Exception as e:
        raise RuntimeError(f"Failed to open the file '{file_path}': {e}")

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' does not exist in the workbook.")

    ws = wb[sheet_name]

    # Get headers from the first row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Map headers to column indexes
    header_index_map = {header: idx for idx, header in enumerate(headers)}

    # Validate requested columns
    if columns:
        for col in columns:
            if col not in header_index_map:
                raise ValueError(f"Column '{col}' not found in the header row.")
        column_indexes = [header_index_map[col] for col in columns]
    else:
        column_indexes = list(range(len(headers)))

    # Validate condition columns
    if condition:
        for col in condition:
            if col not in header_index_map:
                raise ValueError(f"Condition column '{col}' not found in the header row.")
        condition_indexes = {header_index_map[k]: v for k, v in condition.items()}
    else:
        condition_indexes = {}

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Apply condition filtering
        if all(row[idx] == val for idx, val in condition_indexes.items()):
            filtered_row = [row[i] for i in column_indexes]
            data.append(filtered_row)

    return data

# Sample Usage

db = "sample.xlsx"
sheet = "Sample_Sheet"

read_data = opxl_read(db, sheet)
for row in read_data:
    print(row)
    
read_data_fixed_columns = opxl_read(db, sheet, condition={"First Name":"John", "Gender": "Male"}, columns=["Age", "Address"])
for row in read_data_fixed_columns:
    print(row)