import openpyxl

def opxl_update(file_path, sheet_name, new_values, row=None, condition=None):
    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file_path}' does not exist.")
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in workbook.")
    
    ws = wb[sheet_name]

    # Validate new_values is a dict
    if not isinstance(new_values, dict) or not new_values:
        raise ValueError("'new_values' must be a non-empty dictionary.")

    # Map headers to column indices
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    # Check if all keys in new_values exist in headers
    for key in new_values.keys():
        if key not in headers:
            raise ValueError(f"Column '{key}' not found in sheet headers.")

    # If no row or condition provided, update all data rows
    if row is None and condition is None:
        for row_idx in range(2, ws.max_row + 1):
            for col_name, new_val in new_values.items():
                col_idx = headers[col_name]
                ws.cell(row=row_idx, column=col_idx).value = new_val
        
        wb.save(file_path)
        print(f"Updated all rows with new values.")
        return

    # Validate inputs if row or condition is provided
    if row is not None and condition is not None:
        raise ValueError("Provide only one of 'row' or 'condition'.")
    
    if row is not None:
        max_data_row = ws.max_row - 1  # excluding header
        if row < 1 or row > max_data_row:
            raise ValueError(f"Data row {row} is out of range. Max data rows: {max_data_row}")
        excel_row = row + 1
        for col_name, new_val in new_values.items():
            col_idx = headers[col_name]
            ws.cell(row=excel_row, column=col_idx).value = new_val

        wb.save(file_path)
        print(f"Row {row} successfully updated.")
        return

    if condition is not None:
        for key in condition.keys():
            if key not in headers:
                raise ValueError(f"Column '{key}' not found in sheet headers.")

        rows_updated = 0
        for row_idx in range(2, ws.max_row + 1):
            match = True
            for col_name, cond_val in condition.items():
                cell_value = ws.cell(row=row_idx, column=headers[col_name]).value
                if callable(cond_val):
                    if not cond_val(cell_value):
                        match = False
                        break
                else:
                    if cell_value != cond_val:
                        match = False
                        break
            
            if match:
                for col_name, new_val in new_values.items():
                    col_idx = headers[col_name]
                    ws.cell(row=row_idx, column=col_idx).value = new_val
                rows_updated += 1

        if rows_updated == 0:
            print("No rows matched the condition. Nothing updated.")
        else:
            wb.save(file_path)
            print(f"Updated {rows_updated} row(s) matching the condition.")


# Sample Usage

db = "sample.xlsx"
sheet = "Sample_Sheet"

# Update row 2: set "Age" to 28 and "First Name" to "John"
opxl_update(db, sheet, new_values={"Age": 28, "First Name": "John"}, row=2)

# Update all rows where "First Name" == "Sarah" to set "Age" to 30
opxl_update(db, sheet, new_values={"Age": 30}, condition={"First Name": "Sarah"})

# Update all rows where Age <= 17, set "Status" to "Minor"
opxl_update(db, sheet, new_values={"Status": "Minor"}, condition={"Age": lambda x: x is not None and x <= 17})

# Update every row's Status to "Active"
opxl_update(db, sheet, new_values={"Status": "Active"})
