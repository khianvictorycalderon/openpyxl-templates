import os
import openpyxl

def ensure_excel_exist(file_path):
    # Extract the folder path
    directory = os.path.dirname(file_path)

    # Create folders if they do not exist
    if directory and not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory created: {directory}")

    # Create the Excel file if it doesn't exist
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print(f"Excel file created: {file_path}")
    else:
        print(f"Excel file already exists: {file_path}")

def ensure_sheet_exist(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        workbook.save(file_path)
        print(f"Sheet '{sheet_name}' created in {file_path}")
    else:
        print(f"Sheet '{sheet_name}' already exists in {file_path}")

def remove_sheet(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        workbook.save(file_path)
        print(f"Sheet '{sheet_name}' removed from {file_path}")
    else:
        print(f"Sheet '{sheet_name}' does not exist in {file_path}")

# Sample usage:

db = "sample.xlsx"
sheet = "Sample_Sheet"
default_sheet = "Sheet"

ensure_excel_exist(db)
ensure_sheet_exist(db, sheet)
remove_sheet(db, default_sheet)  # Remove default sheet