import openpyxl

def opxl_write_headers(file_path, sheet_name, headers):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        return f"Error: The file '{file_path}' does not exist."

    try:
        ws = wb[sheet_name]
    except KeyError:
        return f"Error: The sheet '{sheet_name}' does not exist in the workbook."

    # Clear existing headers (row 1 only)
    max_column = ws.max_column
    for col in range(1, max_column + 1):
        ws.cell(row=1, column=col).value = None

    # Write new headers
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

    wb.save(file_path)
    return "Headers updated successfully."


# Sample Usage
db = "sample.xlsx"
sheet = "Sample_Sheet"
headers = ["First Name", "Last Name", "Gender", "Age", "Address"]

result = opxl_write_headers(db, sheet, headers)
print(result)