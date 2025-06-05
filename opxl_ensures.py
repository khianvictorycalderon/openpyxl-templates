import os
import openpyxl

def ensure_excel_exist(file_path):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

def ensure_sheet_exist(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
        workbook.save(file_path)

def remove_sheet(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        workbook.save(file_path)

# Sample usage:

db = "sample.xlsx"
sheet = "Sample_Sheet"
default_sheet = "Sheet"

ensure_excel_exist(db)
ensure_sheet_exist(db, sheet)
remove_sheet(db, default_sheet) # Remove default sheet