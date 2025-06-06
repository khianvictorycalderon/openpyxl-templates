import openpyxl
import random
import string

def generate_unique_id_across_sheets(file_path, sheets_name, column=1, id_length=9):
    # Load workbook
    wb = openpyxl.load_workbook(file_path, read_only=True)

    # Collect all existing IDs from given sheets and column
    existing_ids = set()
    for sheet_name in sheets_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_col=column, max_col=column):
                cell_value = row[0].value
                if isinstance(cell_value, str):
                    existing_ids.add(cell_value.strip())

    def generate_id(length):
        chars = string.ascii_uppercase + string.digits
        return "".join(random.choices(chars, k=length))

    # Generate unique ID not in existing_ids
    while True:
        new_id = generate_id(id_length)
        if new_id not in existing_ids:
            return new_id

# Example usage:
print(generate_unique_id_across_sheets("sample.xlsx", ["Sheet 1", "Sheet 2"]))
